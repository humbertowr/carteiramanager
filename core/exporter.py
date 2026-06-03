from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.carteira_processor import obter_consolidacao
from core.formatters import formatar_numero


def exportar_excel_completo(caminho, state):
    if state.df_original is None:
        raise ValueError("Nenhum arquivo importado.")

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        state.gerar_df_prog2().to_excel(writer, sheet_name="PROG 2", index=False)
        state.gerar_df_pedidos_ajustados().to_excel(writer, sheet_name="Pedidos Ajustados", index=False)
        state.gerar_df_pedidos_liberados().to_excel(writer, sheet_name="Pedidos Liberados", index=False)
        state.gerar_df_bloqueios().to_excel(writer, sheet_name="Bloqueios", index=False)
        state.gerar_df_carteira_com_bloqueios().to_excel(writer, sheet_name="Carteira Detalhada", index=False)

        obter_consolidacao(state.df_original, "Por Item").to_excel(writer, sheet_name="Por Item", index=False)
        obter_consolidacao(state.df_original, "Por Pedido").to_excel(writer, sheet_name="Por Pedido", index=False)
        obter_consolidacao(state.df_original, "Por Cliente").to_excel(writer, sheet_name="Por Cliente", index=False)
        obter_consolidacao(state.df_original, "Por Grupo de Faturamento").to_excel(writer, sheet_name="Por Grupo", index=False)
        obter_consolidacao(state.df_original, "Por Previsão de Faturamento").to_excel(writer, sheet_name="Por Previsao", index=False)


def exportar_csv(caminho, df):
    if df is None:
        raise ValueError("Nenhum dado para exportar.")

    df.to_csv(
        caminho,
        index=False,
        sep=";",
        encoding="utf-8-sig"
    )


def preparar_dataframe_exportacao(df, identificador):
    df_saida = df.copy()
    identificador = str(identificador).lower()

    if (
        "itens liberados do prog 2" in identificador
        or "prog2_itens_liberados" in identificador
        or "itens_liberados" in identificador
    ):
        colunas_remover = [
            "Qtd. Pedidos",
            "Qtd. Clientes",
            "Valor Liberado",
            "VLR. ITEM",
        ]

        df_saida.drop(
            columns=[coluna for coluna in colunas_remover if coluna in df_saida.columns],
            inplace=True
        )

        ordem_preferida = [
            "Item",
            "Descrição Item",
            "Qtde Liberada",
        ]

        colunas_existentes = [
            coluna for coluna in ordem_preferida
            if coluna in df_saida.columns
        ]

        outras_colunas = [
            coluna for coluna in df_saida.columns
            if coluna not in colunas_existentes
        ]

        df_saida = df_saida[colunas_existentes + outras_colunas]

    elif (
        "pedidos liberados do prog 2" in identificador
        or "prog2_pedidos_liberados" in identificador
        or "pedidos_liberados" in identificador
    ):
        colunas_remover = [
            "Cliente",
            "Qtd. Itens Liberados",
            "Valor Total Liberado",
            "VLR. PEDIDO",
        ]

        df_saida.drop(
            columns=[coluna for coluna in colunas_remover if coluna in df_saida.columns],
            inplace=True
        )

        ordem_preferida = [
            "Pedido",
            "Grupo",
            "Cliente Original",
            "Data Entrega",
            "Qtde Liberada",
        ]

        colunas_existentes = [
            coluna for coluna in ordem_preferida
            if coluna in df_saida.columns
        ]

        outras_colunas = [
            coluna for coluna in df_saida.columns
            if coluna not in colunas_existentes
        ]

        df_saida = df_saida[colunas_existentes + outras_colunas]

    return df_saida


def exportar_dataframe_excel(caminho, df, sheet_name="Dados"):
    if df is None or df.empty:
        raise ValueError("Nenhum dado para exportar.")

    df_exportar = preparar_dataframe_exportacao(df, sheet_name)

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        df_exportar.to_excel(
            writer,
            sheet_name=sheet_name[:31],
            index=False
        )

        worksheet = writer.sheets[sheet_name[:31]]

        borda_fina = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        formato_moeda = '"R$" #,##0.00'
        colunas_moeda = set()

        for cell in worksheet[1]:
            if cell.value not in (None, ""):
                cell.font = Font(bold=True)
                cell.border = borda_fina

                cabecalho = str(cell.value).upper()
                if "VALOR" in cabecalho or cabecalho.startswith("VLR"):
                    colunas_moeda.add(cell.column)

        for row in worksheet.iter_rows(min_row=2, max_col=worksheet.max_column):
            linha_tem_conteudo = any(cell.value not in (None, "") for cell in row)
            if not linha_tem_conteudo:
                continue

            linha_total = str(row[0].value or "").strip().upper().startswith("TOTAL")

            for cell in row:
                cell.border = borda_fina

                if linha_total:
                    cell.font = Font(bold=True)

                if cell.column in colunas_moeda and cell.value not in (None, ""):
                    cell.number_format = formato_moeda

        for coluna_cells in worksheet.columns:
            tamanho_maximo = 0
            letra_coluna = coluna_cells[0].column_letter

            for cell in coluna_cells:
                valor = "" if cell.value is None else str(cell.value)
                tamanho_maximo = max(tamanho_maximo, len(valor))

            worksheet.column_dimensions[letra_coluna].width = min(tamanho_maximo + 2, 50)


def _valor_numerico(valor):
    try:
        if valor in (None, ""):
            return 0.0
        return float(valor)
    except (TypeError, ValueError):
        return 0.0


def _limpar_texto_exportacao(valor):
    if valor is None or pd.isna(valor):
        return ""
    return str(valor).strip()


def _preparar_faturados_dia_comercial(df):
    if df is None or df.empty:
        raise ValueError("Nenhum dado para exportar.")

    dados = df.copy()

    if "Pedido" in dados.columns:
        dados = dados[~dados["Pedido"].astype(str).str.upper().str.startswith("TOTAL")].copy()

    colunas_origem = {
        "Pedido": "Pedido",
        "Cliente": "Cliente",
        "Item": "Item",
        "Descrição Item": "Descrição do Item",
        "Qtde": "Qtd. Faturada",
        "Valor Total Faturamento": "Valor Faturado",
        "Valor Saldo Pedido": "Saldo do Pedido",
        "Previsão de Embarque": "Previsão de Embarque",
        "OBS": "OBS",
    }

    for coluna in colunas_origem:
        if coluna not in dados.columns:
            dados[coluna] = 0 if coluna in ("Qtde", "Valor Total Faturamento", "Valor Saldo Pedido") else ""

    dados["Qtde"] = pd.to_numeric(dados["Qtde"], errors="coerce").fillna(0)
    dados["Valor Total Faturamento"] = pd.to_numeric(dados["Valor Total Faturamento"], errors="coerce").fillna(0)
    dados["Valor Saldo Pedido"] = pd.to_numeric(dados["Valor Saldo Pedido"], errors="coerce").fillna(0)

    dados.sort_values(["Pedido", "Item"], inplace=True, kind="stable")

    exportar = pd.DataFrame({destino: dados[origem] for origem, destino in colunas_origem.items()})
    exportar["Status"] = dados["Valor Saldo Pedido"].apply(lambda valor: "Parcial" if _valor_numerico(valor) > 0 else "Total")

    ordem = [
        "Pedido",
        "Cliente",
        "Item",
        "Descrição do Item",
        "Qtd. Faturada",
        "Valor Faturado",
        "Saldo do Pedido",
        "Status",
        "Previsão de Embarque",
        "OBS",
    ]
    exportar = exportar[ordem]
    exportar["Saldo do Pedido"] = exportar["Saldo do Pedido"].astype(object)

    # Evita interpretação incorreta do saldo quando o pedido tem mais de um item.
    # O valor fica visível apenas na primeira linha do pedido e o total usa pedido único.
    pedidos_vistos = set()
    for indice, linha in exportar.iterrows():
        pedido = _limpar_texto_exportacao(linha.get("Pedido", ""))
        if not pedido:
            continue
        if pedido in pedidos_vistos:
            exportar.at[indice, "Saldo do Pedido"] = ""
        else:
            pedidos_vistos.add(pedido)

    resumo = {
        "pedidos": dados["Pedido"].astype(str).nunique(),
        "itens": len(dados),
        "quantidade": float(dados["Qtde"].sum()),
        "valor_faturado": float(dados["Valor Total Faturamento"].sum()),
        "saldo_pedido": float(
            dados.drop_duplicates("Pedido")["Valor Saldo Pedido"].sum()
            if "Pedido" in dados.columns else dados["Valor Saldo Pedido"].sum()
        ),
    }

    return exportar, resumo


def exportar_faturados_dia_excel(caminho, df):
    """Exporta os faturados do dia em layout comercial e gerencial."""
    df_exportar, resumo = _preparar_faturados_dia_comercial(df)

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        # Escreve a tabela abaixo do cabeçalho/resumo.
        linha_inicio = 7
        df_exportar.to_excel(
            writer,
            sheet_name="Faturados do Dia",
            index=False,
            startrow=linha_inicio - 1,
        )

        worksheet = writer.sheets["Faturados do Dia"]

        borda_fina = Border(
            left=Side(style="thin", color="D0D7DE"),
            right=Side(style="thin", color="D0D7DE"),
            top=Side(style="thin", color="D0D7DE"),
            bottom=Side(style="thin", color="D0D7DE"),
        )
        borda_titulo = Border(bottom=Side(style="medium", color="1F2937"))
        fill_titulo = PatternFill("solid", fgColor="F8FAFC")
        fill_cabecalho = PatternFill("solid", fgColor="E5E7EB")
        fill_total = PatternFill("solid", fgColor="F3F4F6")
        fonte_titulo = Font(bold=True, size=14, color="111827")
        fonte_negrito = Font(bold=True, color="111827")
        alinhamento_centro = Alignment(horizontal="center", vertical="center")
        alinhamento_texto = Alignment(horizontal="left", vertical="center")
        alinhamento_numero = Alignment(horizontal="right", vertical="center")
        formato_moeda = '"R$" #,##0.00'
        formato_numero = '#,##0.00'

        ultima_coluna = df_exportar.shape[1]
        ultima_linha_tabela = linha_inicio + len(df_exportar)
        linha_total = ultima_linha_tabela + 1

        # Título e resumo.
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ultima_coluna)
        titulo = worksheet.cell(row=1, column=1, value="FATURADOS DO DIA")
        titulo.font = fonte_titulo
        titulo.alignment = alinhamento_texto
        titulo.fill = fill_titulo
        titulo.border = borda_titulo

        resumo_linhas = [
            ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")),
            ("Valor faturado", resumo["valor_faturado"]),
            ("Saldo não faturado", resumo["saldo_pedido"]),
        ]

        for deslocamento, (label, valor) in enumerate(resumo_linhas, start=3):
            worksheet.cell(row=deslocamento, column=1, value=label).font = fonte_negrito
            worksheet.cell(row=deslocamento, column=2, value=valor)

            for coluna in (1, 2):
                cell = worksheet.cell(row=deslocamento, column=coluna)
                cell.border = borda_fina
                cell.alignment = alinhamento_texto if coluna == 1 else alinhamento_numero

        for celula in ("B4", "B5"):
            worksheet[celula].number_format = formato_moeda

        # Cabeçalho da tabela.
        for cell in worksheet[linha_inicio]:
            cell.font = fonte_negrito
            cell.fill = fill_cabecalho
            cell.border = borda_fina
            cell.alignment = alinhamento_centro

        colunas = {cell.value: cell.column for cell in worksheet[linha_inicio]}
        colunas_moeda = {
            colunas.get("Valor Faturado"),
            colunas.get("Saldo do Pedido"),
        }
        colunas_moeda.discard(None)
        coluna_quantidade = colunas.get("Qtd. Faturada")

        # Corpo da tabela.
        for row in worksheet.iter_rows(
            min_row=linha_inicio + 1,
            max_row=ultima_linha_tabela,
            max_col=ultima_coluna,
        ):
            linha_tem_conteudo = any(cell.value not in (None, "") for cell in row)
            if not linha_tem_conteudo:
                continue

            for cell in row:
                cell.border = borda_fina
                cell.alignment = alinhamento_texto

                if cell.column in colunas_moeda and cell.value not in (None, ""):
                    cell.number_format = formato_moeda
                    cell.alignment = alinhamento_numero
                elif cell.column == coluna_quantidade and cell.value not in (None, ""):
                    cell.number_format = formato_numero
                    cell.alignment = alinhamento_numero

        # Total do dia.
        total = {coluna: "" for coluna in df_exportar.columns}
        total["Pedido"] = "TOTAL DO DIA"
        total["Qtd. Faturada"] = resumo["quantidade"]
        total["Valor Faturado"] = resumo["valor_faturado"]
        total["Saldo do Pedido"] = resumo["saldo_pedido"]

        for coluna_idx, coluna_nome in enumerate(df_exportar.columns, start=1):
            cell = worksheet.cell(row=linha_total, column=coluna_idx, value=total[coluna_nome])
            cell.font = fonte_negrito
            cell.fill = fill_total
            cell.border = borda_fina
            cell.alignment = alinhamento_texto

            if coluna_nome in ("Valor Faturado", "Saldo do Pedido"):
                cell.number_format = formato_moeda
                cell.alignment = alinhamento_numero
            elif coluna_nome == "Qtd. Faturada":
                cell.number_format = formato_numero
                cell.alignment = alinhamento_numero

        # Filtros, congelamento e largura.
        ultima_letra_coluna = get_column_letter(ultima_coluna)
        worksheet.auto_filter.ref = f"A{linha_inicio}:{ultima_letra_coluna}{linha_total}"
        worksheet.freeze_panes = f"A{linha_inicio + 1}"

        larguras = {
            "A": 14,  # Pedido
            "B": 38,  # Cliente
            "C": 14,  # Item
            "D": 46,  # Descrição
            "E": 14,  # Qtde
            "F": 18,  # Valor faturado
            "G": 18,  # Saldo
            "H": 12,  # Status
            "I": 22,  # Previsão
            "J": 34,  # OBS
        }

        for letra, largura in larguras.items():
            worksheet.column_dimensions[letra].width = largura

        for row in worksheet.iter_rows(min_row=1, max_row=linha_total, max_col=ultima_coluna):
            for cell in row:
                if cell.value not in (None, ""):
                    cell.alignment = cell.alignment.copy(wrap_text=True)

def preparar_dataframe_pdf(df, titulo):
    df_pdf = preparar_dataframe_exportacao(df, titulo)
    titulo_normalizado = str(titulo).lower()

    if "itens liberados do prog 2" in titulo_normalizado:
        df_pdf.rename(
            columns={
                "Item": "ITEM",
                "Descrição Item": "DESCRICAO ITEM",
                "Qtde Liberada": "QTDE",
            },
            inplace=True
        )

    elif "pedidos liberados do prog 2" in titulo_normalizado:
        df_pdf.rename(
            columns={
                "Pedido": "PEDIDO",
                "Grupo": "GF",
                "Cliente Original": "DESCRICAO CLIENTE",
                "Data Entrega": "ENTREGA",
                "Qtde Liberada": "QTDE",
            },
            inplace=True
        )

    return df_pdf


def titulo_pdf_formatado(titulo):
    titulo_normalizado = str(titulo).lower()

    if "itens liberados do prog 2" in titulo_normalizado:
        return "ITENS PRIORIDADE SEPARACAO"

    if "pedidos liberados do prog 2" in titulo_normalizado:
        return "PEDIDOS PRIORIDADE SEPARACAO"

    return str(titulo).upper()


def formatar_valor_pdf(valor, coluna):
    if pd.isna(valor):
        return ""

    if isinstance(valor, float):
        return formatar_numero(valor)

    if isinstance(valor, int):
        return str(valor)

    return str(valor)


def peso_coluna_pdf(coluna):
    coluna_normalizada = str(coluna).upper()

    if coluna_normalizada == "ITEM":
        return 1.25

    if "DESCRICAO ITEM" in coluna_normalizada:
        return 5.50

    if "DESCRICAO CLIENTE" in coluna_normalizada:
        return 4.30

    if coluna_normalizada == "PEDIDO":
        return 1.10

    if coluna_normalizada == "GF":
        return 0.55

    if "ENTREGA" in coluna_normalizada:
        return 1.10

    if "QTDE" in coluna_normalizada or "QTD" in coluna_normalizada:
        return 0.95

    return 1.00


def alinhamento_coluna_pdf(coluna):
    coluna_normalizada = str(coluna).upper()

    if "QTDE" in coluna_normalizada or "QTD" in coluna_normalizada:
        return "RIGHT"

    if coluna_normalizada == "GF":
        return "CENTER"

    return "LEFT"


def definir_fonte_pdf(df_pdf, titulo):
    titulo_normalizado = str(titulo).lower()
    quantidade_colunas = len(df_pdf.columns)

    if "itens liberados do prog 2" in titulo_normalizado:
        return 10.6, 12.6

    if quantidade_colunas <= 3:
        return 9.8, 11.8

    if quantidade_colunas <= 5:
        return 8.1, 9.8

    return 7.0, 8.6


def desenhar_rodape_pdf(canvas, doc, texto_rodape):
    canvas.saveState()

    canvas.setFont("Courier", 7)

    largura_pagina, _ = doc.pagesize

    y_rodape = 0.45 * 28.3464567

    canvas.drawRightString(
        largura_pagina - doc.rightMargin,
        y_rodape,
        texto_rodape
    )

    canvas.restoreState()


def exportar_dataframe_pdf(caminho, df, titulo="Relatório"):
    if df is None or df.empty:
        raise ValueError("Nenhum dado para exportar.")

    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError as erro:
        raise ImportError(
            "A biblioteca reportlab não está instalada. Rode: pip install reportlab"
        ) from erro

    df_pdf = preparar_dataframe_pdf(df, titulo)
    data_geracao = datetime.now().strftime("%d/%m/%Y %H:%M")
    texto_rodape = f"Gerado em: {data_geracao}"

    estilos = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle(
        "TituloRelatorio",
        parent=estilos["Title"],
        alignment=TA_CENTER,
        fontName="Courier-Bold",
        fontSize=15,
        leading=18,
        spaceAfter=5,
    )

    fonte_tabela, leading_tabela = definir_fonte_pdf(df_pdf, titulo)

    estilo_texto = ParagraphStyle(
        "TextoTabela",
        parent=estilos["BodyText"],
        fontName="Courier",
        fontSize=fonte_tabela,
        leading=leading_tabela,
    )

    estilo_cabecalho = ParagraphStyle(
        "CabecalhoTabela",
        parent=estilos["BodyText"],
        fontName="Courier-Bold",
        fontSize=fonte_tabela,
        leading=leading_tabela,
    )

    documento = SimpleDocTemplate(
        caminho,
        pagesize=A4,
        rightMargin=0.6 * cm,
        leftMargin=0.6 * cm,
        topMargin=0.9 * cm,
        bottomMargin=1.0 * cm,
    )

    elementos = [
        Paragraph(titulo_pdf_formatado(titulo), estilo_titulo),
        Spacer(1, 0.12 * cm),
    ]

    colunas = list(df_pdf.columns)

    dados = [
        [Paragraph(str(coluna), estilo_cabecalho) for coluna in colunas]
    ]

    for _, linha in df_pdf.iterrows():
        dados.append([
            Paragraph(formatar_valor_pdf(linha[coluna], coluna), estilo_texto)
            for coluna in colunas
        ])

    largura_total = A4[0] - (1.2 * cm)

    pesos = [
        peso_coluna_pdf(coluna)
        for coluna in colunas
    ]

    soma_pesos = sum(pesos) if pesos else 1

    larguras = [
        largura_total * (peso / soma_pesos)
        for peso in pesos
    ]

    tabela = Table(
        dados,
        colWidths=larguras,
        repeatRows=1,
        splitByRow=True,
    )

    estilos_tabela = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.white),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Courier-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Courier"),
        ("FONTSIZE", (0, 0), (-1, -1), fonte_tabela),
        ("TOPPADDING", (0, 0), (-1, 0), 4),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
        ("TOPPADDING", (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("BOX", (0, 0), (-1, 0), 0.75, colors.black),
        ("LINEBELOW", (0, 0), (-1, 0), 0.75, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]

    for indice, coluna in enumerate(colunas):
        alinhamento = alinhamento_coluna_pdf(coluna)
        estilos_tabela.append(("ALIGN", (indice, 0), (indice, -1), alinhamento))

    tabela.setStyle(TableStyle(estilos_tabela))

    elementos.append(tabela)

    documento.build(
        elementos,
        onFirstPage=lambda canvas, doc: desenhar_rodape_pdf(canvas, doc, texto_rodape),
        onLaterPages=lambda canvas, doc: desenhar_rodape_pdf(canvas, doc, texto_rodape),
    )

def exportar_simulacao_faturamento_excel(caminho, resultado):
    """Exporta a simulação semanal com uma aba de visão geral e uma aba por dia."""
    if not resultado or not resultado.get("dias"):
        raise ValueError("Nenhuma simulação disponível para exportar.")

    from openpyxl import Workbook

    wb = Workbook()
    ws_geral = wb.active
    ws_geral.title = "Visão Geral"

    estilos = _criar_estilos_simulacao()
    _montar_aba_visao_geral_simulacao(ws_geral, resultado, estilos)

    nomes_usados = {ws_geral.title}
    for dia in resultado.get("dias", []):
        nome_aba = _nome_aba_dia_simulacao(dia.get("data"), nomes_usados)
        nomes_usados.add(nome_aba)
        ws_dia = wb.create_sheet(nome_aba)
        _montar_aba_dia_simulacao(ws_dia, dia, estilos)

    wb.save(caminho)


def _criar_estilos_simulacao():
    return {
        "border": Border(
            left=Side(style="thin", color="D0D7DE"),
            right=Side(style="thin", color="D0D7DE"),
            top=Side(style="thin", color="D0D7DE"),
            bottom=Side(style="thin", color="D0D7DE"),
        ),
        "fill_title": PatternFill("solid", fgColor="F8FAFC"),
        "fill_header": PatternFill("solid", fgColor="E5E7EB"),
        "fill_section": PatternFill("solid", fgColor="F3F4F6"),
        "fill_total": PatternFill("solid", fgColor="EEF2FF"),
        "font_title": Font(bold=True, size=14, color="111827"),
        "font_header": Font(bold=True, color="111827"),
        "font_muted": Font(color="4B5563"),
        "align_left": Alignment(horizontal="left", vertical="center"),
        "align_center": Alignment(horizontal="center", vertical="center"),
        "align_right": Alignment(horizontal="right", vertical="center"),
        "currency": '"R$" #,##0.00',
        "number": '#,##0.00',
        "percent": '0.00%',
    }


def _montar_aba_visao_geral_simulacao(ws, resultado, estilos):
    dias = resultado.get("dias", [])
    periodo = f"{resultado['data_inicio'].strftime('%d/%m/%Y')} a {resultado['data_fim'].strftime('%d/%m/%Y')}"
    total_bloqueado = sum(
        float(pedido.get("valor_bloqueado", 0) or 0)
        for dia in dias
        for pedido in dia.get("pedidos", [])
    )
    total_itens = sum(
        len(pedido.get("itens", []))
        for dia in dias
        for pedido in dia.get("pedidos", [])
    )

    ws.merge_cells("A1:I1")
    ws["A1"] = "SIMULAÇÃO SEMANAL DE FATURAMENTO"
    ws["A1"].font = estilos["font_title"]
    ws["A1"].fill = estilos["fill_title"]
    ws["A1"].alignment = estilos["align_left"]

    resumo = [
        ("Período", periodo, "Dias", resultado.get("qtd_dias", len(dias))),
        ("Meta diária", resultado.get("meta_diaria", 0), "Meta total", resultado.get("meta_total", 0)),
        ("Valor estimado", resultado.get("valor_estimado_total", 0), "Diferença", resultado.get("diferenca_total", 0)),
        ("Pedidos", resultado.get("qtd_pedidos", 0), "Itens", total_itens),
        ("Saldo bloqueado", total_bloqueado, "Saldo elegível restante", resultado.get("valor_restante", 0)),
    ]

    for linha, (label_a, valor_b, label_d, valor_e) in enumerate(resumo, start=3):
        ws.cell(row=linha, column=1, value=label_a).font = estilos["font_header"]
        ws.cell(row=linha, column=2, value=valor_b)
        ws.cell(row=linha, column=4, value=label_d).font = estilos["font_header"]
        ws.cell(row=linha, column=5, value=valor_e)
        for coluna in (1, 2, 4, 5):
            cell = ws.cell(row=linha, column=coluna)
            cell.border = estilos["border"]
            cell.alignment = estilos["align_left"] if coluna in (1, 4) else estilos["align_right"]
            if linha in (4, 5, 7) and coluna in (2, 5):
                cell.number_format = estilos["currency"]

    linha_inicio = 10
    cabecalhos = [
        "Data",
        "Meta",
        "Valor Estimado",
        "Diferença",
        "% da Meta",
        "Pedidos",
        "Itens",
        "Saldo Bloqueado",
        "Status",
    ]
    _escrever_cabecalho(ws, linha_inicio, cabecalhos, estilos)

    for offset, dia in enumerate(dias, start=1):
        linha = linha_inicio + offset
        meta = float(dia.get("meta", 0) or 0)
        valor = float(dia.get("valor_estimado", 0) or 0)
        itens = sum(len(pedido.get("itens", [])) for pedido in dia.get("pedidos", []))
        saldo_bloqueado = sum(float(pedido.get("valor_bloqueado", 0) or 0) for pedido in dia.get("pedidos", []))
        valores = [
            dia.get("data").strftime("%d/%m/%Y") if dia.get("data") else "",
            meta,
            valor,
            float(dia.get("diferenca", 0) or 0),
            (valor / meta) if meta else 0,
            len(dia.get("pedidos", [])),
            itens,
            saldo_bloqueado,
            dia.get("status", ""),
        ]
        for coluna, valor_cell in enumerate(valores, start=1):
            cell = ws.cell(row=linha, column=coluna, value=valor_cell)
            cell.border = estilos["border"]
            cell.alignment = estilos["align_center"] if coluna in (1, 6, 7, 9) else estilos["align_right"]
            if coluna in (2, 3, 4, 8):
                cell.number_format = estilos["currency"]
            elif coluna == 5:
                cell.number_format = estilos["percent"]

    linha_total = linha_inicio + len(dias) + 1
    total_meta = resultado.get("meta_total", 0)
    total_estimado = resultado.get("valor_estimado_total", 0)
    total_diff = resultado.get("diferenca_total", 0)
    total_pedidos = resultado.get("qtd_pedidos", 0)
    total_percentual = (float(total_estimado or 0) / float(total_meta or 1)) if total_meta else 0
    totais = ["TOTAL", total_meta, total_estimado, total_diff, total_percentual, total_pedidos, total_itens, total_bloqueado, ""]
    _escrever_linha_total(ws, linha_total, totais, estilos, colunas_moeda={2, 3, 4, 8}, colunas_percentual={5})

    ws.auto_filter.ref = f"A{linha_inicio}:I{linha_total}"
    ws.freeze_panes = f"A{linha_inicio + 1}"
    _ajustar_layout_planilha(ws, {
        "A": 13,
        "B": 15,
        "C": 18,
        "D": 15,
        "E": 12,
        "F": 10,
        "G": 10,
        "H": 18,
        "I": 22,
    })


def _montar_aba_dia_simulacao(ws, dia, estilos):
    data_texto = dia.get("data").strftime("%d/%m/%Y") if dia.get("data") else ""
    pedidos = dia.get("pedidos", [])
    saldo_bloqueado_total = sum(float(pedido.get("valor_bloqueado", 0) or 0) for pedido in pedidos)

    ws.merge_cells("A1:G1")
    ws["A1"] = f"SIMULAÇÃO DE FATURAMENTO - {data_texto}"
    ws["A1"].font = estilos["font_title"]
    ws["A1"].fill = estilos["fill_title"]
    ws["A1"].alignment = estilos["align_left"]

    linha_inicio = 3
    cabecalhos = [
        "Pedido",
        "Item",
        "Descrição do Item",
        "Cliente",
        "QTD",
        "Valor Faturável",
        "Saldo Bloqueado",
    ]
    _escrever_cabecalho(ws, linha_inicio, cabecalhos, estilos)

    linha_atual = linha_inicio + 1
    if not pedidos:
        ws.cell(row=linha_atual, column=1, value="Nenhum pedido selecionado para este dia.")
        ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=len(cabecalhos))
        ws.cell(row=linha_atual, column=1).font = estilos["font_muted"]
        ws.cell(row=linha_atual, column=1).alignment = estilos["align_left"]
        for col in range(1, len(cabecalhos) + 1):
            ws.cell(row=linha_atual, column=col).border = estilos["border"]
        linha_atual += 1
    else:
        for pedido in pedidos:
            itens = pedido.get("itens", [])
            if not itens:
                valores = [
                    pedido.get("pedido", ""),
                    "",
                    "Sem itens detalhados",
                    pedido.get("cliente_original") or pedido.get("cliente", ""),
                    pedido.get("quantidade", ""),
                    pedido.get("valor_liberado", 0),
                    pedido.get("valor_bloqueado", 0),
                ]
                _escrever_linha_dia_simulacao(ws, linha_atual, valores, estilos, linha_pedido=True)
                linha_atual += 1
                continue

            primeiro_item = True
            for item in itens:
                valores = [
                    pedido.get("pedido", "") if primeiro_item else "",
                    item.get("item", ""),
                    item.get("descricao", ""),
                    pedido.get("cliente_original") or pedido.get("cliente", "") if primeiro_item else "",
                    item.get("quantidade", ""),
                    item.get("valor", ""),
                    pedido.get("valor_bloqueado", 0) if primeiro_item else "",
                ]
                _escrever_linha_dia_simulacao(ws, linha_atual, valores, estilos, linha_pedido=primeiro_item)
                linha_atual += 1
                primeiro_item = False

    linha_total = linha_atual
    totais = ["TOTAL", "", "", "", "", dia.get("valor_estimado", 0), saldo_bloqueado_total]
    _escrever_linha_total(ws, linha_total, totais, estilos, colunas_moeda={6, 7})

    ws.auto_filter.ref = f"A{linha_inicio}:G{linha_total}"
    ws.freeze_panes = f"A{linha_inicio + 1}"
    _ajustar_layout_planilha(ws, {
        "A": 15,
        "B": 16,
        "C": 48,
        "D": 42,
        "E": 12,
        "F": 18,
        "G": 18,
    })


def _escrever_cabecalho(ws, linha, cabecalhos, estilos):
    for coluna, titulo in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=linha, column=coluna, value=titulo)
        cell.font = estilos["font_header"]
        cell.fill = estilos["fill_header"]
        cell.border = estilos["border"]
        cell.alignment = estilos["align_center"]


def _escrever_linha_dia_simulacao(ws, linha, valores, estilos, linha_pedido=False):
    for coluna, valor in enumerate(valores, start=1):
        cell = ws.cell(row=linha, column=coluna, value=valor)
        cell.border = estilos["border"]
        cell.alignment = estilos["align_left"]
        if linha_pedido:
            cell.fill = estilos["fill_section"]
            if coluna in (1, 4):
                cell.font = estilos["font_header"]
        if coluna in (5, 6, 7) and valor not in (None, ""):
            cell.alignment = estilos["align_right"]
            cell.number_format = estilos["number"] if coluna == 5 else estilos["currency"]


def _escrever_linha_total(ws, linha, valores, estilos, colunas_moeda=None, colunas_percentual=None):
    colunas_moeda = colunas_moeda or set()
    colunas_percentual = colunas_percentual or set()
    for coluna, valor in enumerate(valores, start=1):
        cell = ws.cell(row=linha, column=coluna, value=valor)
        cell.font = estilos["font_header"]
        cell.fill = estilos["fill_total"]
        cell.border = estilos["border"]
        cell.alignment = estilos["align_right"] if coluna != 1 else estilos["align_left"]
        if coluna in colunas_moeda and valor not in (None, ""):
            cell.number_format = estilos["currency"]
        elif coluna in colunas_percentual and valor not in (None, ""):
            cell.number_format = estilos["percent"]


def _ajustar_layout_planilha(ws, larguras):
    for letra, largura in larguras.items():
        ws.column_dimensions[letra].width = largura

    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                cell.alignment = cell.alignment.copy(wrap_text=True)

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 22
    ws.row_dimensions[1].height = 26


def _nome_aba_dia_simulacao(data_dia, nomes_usados):
    if data_dia:
        nome_base = data_dia.strftime("%d-%m")
    else:
        nome_base = "Dia"

    nome = nome_base[:31]
    contador = 2
    while nome in nomes_usados:
        sufixo = f"_{contador}"
        nome = f"{nome_base[:31 - len(sufixo)]}{sufixo}"
        contador += 1
    return nome
